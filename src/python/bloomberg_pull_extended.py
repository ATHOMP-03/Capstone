"""
bloomberg_pull_extended.py
Bloomberg Terminal data pull for the Teti et al. (2019) and Gu & Kurov (2020)
replication studies. Run this script on a machine with an active Bloomberg
Terminal session (desktop Bloomberg app must be open and logged in).

Fetches all fields needed by both replication scripts:

  EXISTING FIELDS (already in panel_long.csv — included here for a clean
  2024-2026 re-pull with one consolidated file):
    PX_OPEN, PX_OFFICIAL_CLOSE, PX_HIGH, PX_LOW
    CUR_MKT_CAP, TOTAL_EQUITY, TOT_DEBT_TO_TOT_EQY
    PX_VOLUME
    TWITTER_SENTIMENT_DAILY_AVG, TWITTER_PUBLICATION_COUNT
    TWITTER_NEG_SENTIMENT_COUNT
    NEWS_SENTIMENT_DAILY_AVG
    RSI_30D, MOV_AVG_50D

  NEW FIELDS (required for Teti and/or Gu & Kurov):
    TWITTER_POS_SENTIMENT_COUNT   [Teti SET 2 & 3 — polarity-broken tweet counts]
    TWITTER_NEU_SENTIMENT_COUNT   [Teti SET 2 & 3 — polarity-broken tweet counts]
    BID_ASK_SPREAD_DAILY_AVG      [Gu & Kurov — daily average bid-ask spread in $]

  REFERENCE DATA (point-in-time, not historical — pulled separately):
    TWITTER_FOLLOWERS             [Teti SET 2 & 3 — group dummy construction]
    NUM_ANALYST                   [Gu & Kurov Table 4 — analyst coverage]

Requirements:
    pip install blpapi pandas openpyxl
    Bloomberg Desktop API must be running (blpapi uses port 8194 by default).
    Install blpapi wheel from https://bloomberg.com/professional/support/api-library/

Outputs:
    data/raw/bloomberg_2024_2026.csv      — daily panel (long format, all tickers)
    data/raw/bloomberg_ref_data.csv       — reference data (followers, analyst count)
    data/processed/panel_long_extended.csv — cleaned, merged, analysis-ready panel

Usage:
    python src/python/bloomberg_pull_extended.py

Bloomberg Terminal verification:
    To confirm a field name, type in the terminal:
        FLDS <field_name> <GO>
    e.g., FLDS TWITTER_POS_SENTIMENT_COUNT <GO>
         FLDS BID_ASK_SPREAD_DAILY_AVG <GO>
"""

from __future__ import annotations

import sys
import warnings
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

ROOT     = Path(__file__).resolve().parents[2]
RAW_DIR  = ROOT / "data" / "raw"
PROC_DIR = ROOT / "data" / "processed"
RAW_DIR.mkdir(parents=True, exist_ok=True)
PROC_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Configuration — update dates and tickers as needed
# ---------------------------------------------------------------------------
START_DATE = date(2024, 1, 1)
END_DATE   = date(2026, 3, 31)   # extend as your data goes

# Tickers: loaded from existing panel_long.csv if available,
# otherwise define manually as Bloomberg identifiers.
EXISTING_PANEL = PROC_DIR / "panel_long.csv"
if EXISTING_PANEL.exists():
    _existing = pd.read_csv(EXISTING_PANEL, usecols=["ticker"])
    TICKERS = sorted(_existing["ticker"].dropna().unique().tolist())
    print(f"Loaded {len(TICKERS)} tickers from existing panel_long.csv")
else:
    # Fallback: define your Bloomberg security identifiers here
    TICKERS = [
        "AAPL US Equity",
        "MSFT US Equity",
        "GOOGL US Equity",
        "META US Equity",
        "NVDA US Equity",
        # Add all tickers from your study here
    ]
    print(f"Using {len(TICKERS)} manually defined tickers")

# ---------------------------------------------------------------------------
# Field definitions
# ---------------------------------------------------------------------------
HISTORICAL_FIELDS = [
    # --- Price fields ---
    "PX_OPEN",
    "PX_OFFICIAL_CLOSE",
    "PX_HIGH",
    "PX_LOW",
    # --- Fundamentals ---
    "CUR_MKT_CAP",
    "TOTAL_EQUITY",
    "TOT_DEBT_TO_TOT_EQY",
    "PX_VOLUME",
    # --- Twitter sentiment (Bloomberg/Twitter Data Agreement) ---
    "TWITTER_SENTIMENT_DAILY_AVG",
    "TWITTER_PUBLICATION_COUNT",
    "TWITTER_NEG_SENTIMENT_COUNT",
    "TWITTER_POS_SENTIMENT_COUNT",      # NEW: required for Teti SET 2 & 3
    "TWITTER_NEU_SENTIMENT_COUNT",      # NEW: required for Teti SET 2 & 3
    # --- News sentiment ---
    "NEWS_SENTIMENT_DAILY_AVG",
    # --- Technical indicators ---
    "RSI_30D",
    "MOV_AVG_50D",
    # --- Liquidity (Gu & Kurov Table 2-5 control) ---
    "BID_ASK_SPREAD_DAILY_AVG",         # NEW: required for Gu & Kurov
]

REFERENCE_FIELDS = {
    "TWITTER_FOLLOWERS": "twitter_followers",    # Teti group dummy
    "NUM_ANALYST":       "analyst_count",        # Gu & Kurov heterogeneity
}

FIELD_RENAME = {
    "PX_OPEN":                      "px_open",
    "PX_OFFICIAL_CLOSE":            "px_close",
    "PX_HIGH":                      "px_high",
    "PX_LOW":                       "px_low",
    "CUR_MKT_CAP":                  "mkt_cap",
    "TOTAL_EQUITY":                 "total_equity",
    "TOT_DEBT_TO_TOT_EQY":         "debt_to_equity",
    "PX_VOLUME":                    "volume",
    "TWITTER_SENTIMENT_DAILY_AVG":  "twitter_sent",
    "TWITTER_PUBLICATION_COUNT":    "twitter_count",
    "TWITTER_NEG_SENTIMENT_COUNT":  "twitter_neg_count",
    "TWITTER_POS_SENTIMENT_COUNT":  "twitter_pos_count",
    "TWITTER_NEU_SENTIMENT_COUNT":  "twitter_neu_count",
    "NEWS_SENTIMENT_DAILY_AVG":     "news_sent",
    "RSI_30D":                      "rsi_30",
    "MOV_AVG_50D":                  "ma_50",
    "BID_ASK_SPREAD_DAILY_AVG":     "bid_ask_spread",
}

BLOOMBERG_NA = ["#N/A N/A", "#N/A", "#N/A Field Not Applicable", "N.A.", "N/A"]
BATCH_SIZE   = 25   # Bloomberg recommends ≤50 securities per request


# ===========================================================================
# blpapi HISTORICAL DATA PULL
# ===========================================================================

def _pull_bdh_batch(
    session,
    securities: list[str],
    fields: list[str],
    start: date,
    end: date,
) -> pd.DataFrame:
    """Fetch historical data for a batch of securities."""
    try:
        import blpapi
    except ImportError:
        raise ImportError("blpapi not installed. See: https://bloomberg.com/professional/support/api-library/")

    ref_svc = session.getService("//blp/refdata")
    request = ref_svc.createRequest("HistoricalDataRequest")

    for sec in securities:
        request.getElement("securities").appendValue(sec)
    for fld in fields:
        request.getElement("fields").appendValue(fld)

    request.set("startDate", start.strftime("%Y%m%d"))
    request.set("endDate",   end.strftime("%Y%m%d"))
    request.set("periodicityAdjustment", "ACTUAL")
    request.set("periodicitySelection",  "DAILY")
    request.set("nonTradingDayFillOption", "NIL_VALUE")
    request.set("nonTradingDayFillMethod", "NIL_VALUE")

    session.sendRequest(request)

    rows = []
    done = False
    while not done:
        event = session.nextEvent(2000)
        for msg in event:
            msg_type = str(msg.messageType())
            if msg_type in ("HistoricalDataResponse",):
                sec_data   = msg.getElement("securityData")
                security   = sec_data.getElementAsString("security")
                field_data = sec_data.getElement("fieldData")
                for i in range(field_data.numValues()):
                    record = field_data.getValue(i)
                    row = {
                        "date":   record.getElementAsDatetime("date").date(),
                        "ticker": security,
                    }
                    for fld in fields:
                        try:
                            row[fld] = record.getElementAsFloat(fld)
                        except Exception:
                            row[fld] = np.nan
                    rows.append(row)
        import blpapi as _blp
        if event.eventType() == _blp.Event.RESPONSE:
            done = True

    return pd.DataFrame(rows)


def _pull_bdp_batch(
    session,
    securities: list[str],
    fields: dict[str, str],  # bloomberg_field -> snake_case_name
) -> pd.DataFrame:
    """Fetch reference (point-in-time) data for a batch of securities."""
    try:
        import blpapi
    except ImportError:
        raise ImportError("blpapi not installed.")

    ref_svc = session.getService("//blp/refdata")
    request = ref_svc.createRequest("ReferenceDataRequest")

    for sec in securities:
        request.getElement("securities").appendValue(sec)
    for fld in fields:
        request.getElement("fields").appendValue(fld)

    session.sendRequest(request)

    rows = []
    done = False
    while not done:
        event = session.nextEvent(2000)
        for msg in event:
            msg_type = str(msg.messageType())
            if "ReferenceDataResponse" in msg_type:
                sec_array = msg.getElement("securityData")
                for i in range(sec_array.numValues()):
                    sec_el   = sec_array.getValue(i)
                    security = sec_el.getElementAsString("security")
                    fld_data = sec_el.getElement("fieldData")
                    row = {"ticker": security}
                    for bbg_fld, col_name in fields.items():
                        try:
                            row[col_name] = fld_data.getElementAsFloat(bbg_fld)
                        except Exception:
                            try:
                                row[col_name] = fld_data.getElementAsString(bbg_fld)
                            except Exception:
                                row[col_name] = np.nan
                    rows.append(row)
        import blpapi as _blp
        if event.eventType() == _blp.Event.RESPONSE:
            done = True

    return pd.DataFrame(rows)


def pull_bloomberg_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Pull all required Bloomberg data. Returns (historical_df, reference_df)."""
    try:
        import blpapi
    except ImportError as exc:
        raise ImportError(
            "blpapi package not found.\n"
            "Install from: https://bloomberg.com/professional/support/api-library/\n"
            "Or use the Excel BDH template generated by this script (see fallback below)."
        ) from exc

    session_opts = blpapi.SessionOptions()
    session_opts.setServerHost("localhost")
    session_opts.setServerPort(8194)

    session = blpapi.Session(session_opts)
    if not session.start():
        raise RuntimeError(
            "Failed to connect to Bloomberg Terminal.\n"
            "Ensure Bloomberg Desktop is running and you are logged in."
        )
    if not session.openService("//blp/refdata"):
        raise RuntimeError("Failed to open Bloomberg refdata service.")

    print(f"Connected to Bloomberg Terminal.")
    print(f"Pulling {len(TICKERS)} tickers × {len(HISTORICAL_FIELDS)} fields")
    print(f"Date range: {START_DATE} → {END_DATE}")

    # --- Historical data (BDH) ---
    hist_chunks = []
    n_batches = (len(TICKERS) + BATCH_SIZE - 1) // BATCH_SIZE
    for i in range(0, len(TICKERS), BATCH_SIZE):
        batch = TICKERS[i:i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        print(f"  Batch {batch_num}/{n_batches}: {len(batch)} securities …", end="\r", flush=True)
        chunk = _pull_bdh_batch(session, batch, HISTORICAL_FIELDS, START_DATE, END_DATE)
        hist_chunks.append(chunk)
    print()

    hist_df = pd.concat(hist_chunks, ignore_index=True)

    # --- Reference data (BDP) ---
    ref_chunks = []
    print(f"Pulling reference data: {list(REFERENCE_FIELDS.keys())}")
    for i in range(0, len(TICKERS), BATCH_SIZE):
        batch = TICKERS[i:i + BATCH_SIZE]
        chunk = _pull_bdp_batch(session, batch, REFERENCE_FIELDS)
        ref_chunks.append(chunk)

    ref_df = pd.concat(ref_chunks, ignore_index=True)

    session.stop()
    print("Bloomberg session closed.")

    return hist_df, ref_df


# ===========================================================================
# CLEANING PIPELINE
# ===========================================================================

def clean_and_save(hist_df: pd.DataFrame, ref_df: pd.DataFrame) -> pd.DataFrame:
    """Clean raw Bloomberg output and produce analysis-ready panel."""

    # ----- Historical data -----
    hist_df["date"] = pd.to_datetime(hist_df["date"])

    # Rename Bloomberg field codes → snake_case
    hist_df = hist_df.rename(columns=FIELD_RENAME)

    numeric_cols = list(FIELD_RENAME.values())
    for col in numeric_cols:
        if col in hist_df.columns:
            hist_df[col] = hist_df[col].replace(BLOOMBERG_NA, np.nan)
            hist_df[col] = pd.to_numeric(hist_df[col], errors="coerce")

    # Drop non-trading days
    hist_df = hist_df.dropna(subset=["px_open", "px_close"]).reset_index(drop=True)
    hist_df = hist_df.sort_values(["ticker", "date"]).reset_index(drop=True)

    # Merge reference data onto historical panel (applies to all dates for each ticker)
    if not ref_df.empty:
        for col in ref_df.columns:
            if col in REFERENCE_FIELDS.values():
                ref_df[col] = pd.to_numeric(ref_df[col], errors="coerce")
        hist_df = hist_df.merge(
            ref_df[["ticker"] + list(REFERENCE_FIELDS.values())],
            on="ticker", how="left"
        )

    # ----- Derived variables -----
    hist_df["return"]     = hist_df["px_close"] - hist_df["px_open"]
    hist_df["pct_return"] = (hist_df["px_close"] - hist_df["px_open"]) / hist_df["px_open"] * 100

    for n in [1, 2, 3, 5, 7]:
        hist_df[f"lag{n}"] = hist_df.groupby("ticker")["return"].shift(n)

    # Open-to-open return (Gu & Kurov)
    hist_df["px_open_next"] = hist_df.groupby("ticker")["px_open"].shift(-1)
    hist_df["return_oo"] = (
        (hist_df["px_open_next"] - hist_df["px_open"]) / hist_df["px_open"] * 100
    )

    # Abnormal volume
    mean_vol = hist_df.groupby("ticker")["volume"].transform("mean")
    hist_df["abnorm_vol"] = ((hist_df["volume"] - mean_vol) / mean_vol) * 100

    # Rogers-Satchell volatility
    lH = np.log(hist_df["px_high"].clip(lower=1e-8))
    lL = np.log(hist_df["px_low"].clip(lower=1e-8))
    lO = np.log(hist_df["px_open"].clip(lower=1e-8))
    lC = np.log(hist_df["px_close"].clip(lower=1e-8))
    hist_df["vol_rs"] = ((lH - lC) * (lH - lO) + (lL - lC) * (lL - lO)).clip(lower=0) * 100

    # Log market cap
    hist_df["log_mkt_cap"] = np.log(hist_df["mkt_cap"].clip(lower=1e-8))

    # Polarity-weighted sentiment index ts_b (Teti SET 3)
    if {"twitter_pos_count", "twitter_neu_count", "twitter_neg_count"}.issubset(hist_df.columns):
        denom = (
            hist_df["twitter_pos_count"]
            + hist_df["twitter_neu_count"]
            + hist_df["twitter_neg_count"]
        ).replace(0, np.nan)
        hist_df["ts_b"] = (hist_df["twitter_pos_count"] - hist_df["twitter_neg_count"]) / denom

    # ----- Save -----
    raw_path  = RAW_DIR  / "bloomberg_2024_2026.csv"
    proc_path = PROC_DIR / "panel_long_extended.csv"

    # Save raw (pre-cleaning) separately
    raw_out = hist_df.rename(columns={v: k for k, v in FIELD_RENAME.items() if v in hist_df.columns})
    ref_df.to_csv(RAW_DIR / "bloomberg_ref_data.csv", index=False)
    print(f"Saved {RAW_DIR / 'bloomberg_ref_data.csv'}")

    hist_df.to_csv(proc_path, index=False)
    print(f"Saved {proc_path}  ({len(hist_df):,} rows × {hist_df.shape[1]} cols)")

    return hist_df


# ===========================================================================
# FALLBACK — Generate Bloomberg Excel BDH template
#
# If blpapi is not available, this creates an Excel workbook with BDH formulas
# that will auto-populate when opened in Bloomberg-connected Excel.
# File: data/raw/bloomberg_bdh_template.xlsx
# ===========================================================================

def generate_excel_template():
    """Generate an Excel file with Bloomberg BDH formulas for manual use."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        _print_manual_instructions()
        return

    wb = openpyxl.Workbook()

    # --- Sheet 1: Historical data BDH formulas ---
    ws_hist = wb.active
    ws_hist.title = "Historical_BDH"

    start_str = START_DATE.strftime("%Y%m%d")
    end_str   = END_DATE.strftime("%Y%m%d")

    headers = ["Security", "Field", "Start Date", "End Date",
               "Periodicity", "BDH Formula"]
    ws_hist.append(headers)

    new_fields_only = [
        "TWITTER_POS_SENTIMENT_COUNT",
        "TWITTER_NEU_SENTIMENT_COUNT",
        "BID_ASK_SPREAD_DAILY_AVG",
    ]
    instructions = (
        "INSTRUCTIONS: In Bloomberg Excel, paste each BDH formula into an empty cell. "
        "Data will auto-populate. Then export/copy to the main Bloomberg export file."
    )
    ws_hist.append([instructions])

    for ticker in TICKERS[:5]:   # show first 5 as examples
        for fld in new_fields_only:
            formula = (
                f'=BDH("{ticker}","{fld}","{start_str}","{end_str}",'
                f'"Days=A","Fill=P","Per=D")'
            )
            ws_hist.append([ticker, fld, start_str, end_str, "DAILY", formula])

    # --- Sheet 2: Reference data BDP formulas ---
    ws_ref = wb.create_sheet("Reference_BDP")
    ws_ref.append(["Security", "Field (TWITTER_FOLLOWERS)", "BDP Formula",
                   "Field (NUM_ANALYST)", "BDP Formula"])
    for ticker in TICKERS[:5]:
        ws_ref.append([
            ticker,
            "TWITTER_FOLLOWERS",
            f'=BDP("{ticker}","TWITTER_FOLLOWERS")',
            "NUM_ANALYST",
            f'=BDP("{ticker}","NUM_ANALYST")',
        ])

    # --- Sheet 3: All tickers list ---
    ws_tickers = wb.create_sheet("Ticker_List")
    ws_tickers.append(["Ticker", "Notes"])
    for t in TICKERS:
        ws_tickers.append([t, ""])

    out_path = RAW_DIR / "bloomberg_bdh_template.xlsx"
    wb.save(out_path)
    print(f"\nExcel BDH template saved: {out_path}")
    print("Open this file in Bloomberg-connected Excel to auto-populate data.")
    print("Then copy the data into your standard Bloomberg export format.")


def _print_manual_instructions():
    """Print Bloomberg terminal commands for manual data pull."""
    print("\n" + "="*65)
    print("MANUAL BLOOMBERG TERMINAL INSTRUCTIONS")
    print("="*65)
    print("\nVerify field names (type in terminal):")
    print("  FLDS TWITTER_POS_SENTIMENT_COUNT <GO>")
    print("  FLDS TWITTER_NEU_SENTIMENT_COUNT <GO>")
    print("  FLDS BID_ASK_SPREAD_DAILY_AVG <GO>")
    print("  FLDS TWITTER_FOLLOWERS <GO>")
    print("  FLDS NUM_ANALYST <GO>")
    print("\nDownload via Excel BDH (example for one ticker):")
    print(f"  =BDH(\"AAPL US Equity\",\"TWITTER_POS_SENTIMENT_COUNT\",")
    print(f"        \"{START_DATE.strftime('%Y%m%d')}\",\"{END_DATE.strftime('%Y%m%d')}\",")
    print(f"        \"Days=A\",\"Fill=P\",\"Per=D\")")
    print("\nFor all tickers, use the DAPI Excel template or Bloomberg's:")
    print("  XLTP <GO>  — Excel template picker")
    print("  API <GO>   — Bloomberg API documentation")
    print("  WAPI <GO>  — Bloomberg API getting started guide")
    print("\nAlternatively, add the new fields to your existing")
    print("Bloomberg25MAR.xlsx template and re-run clean_data.py.")


# ===========================================================================
# MAIN
# ===========================================================================

if __name__ == "__main__":
    try:
        import blpapi  # noqa: F401
        print("blpapi detected. Connecting to Bloomberg Terminal…")
        hist_df, ref_df = pull_bloomberg_data()
        panel = clean_and_save(hist_df, ref_df)
        print(f"\nDone. panel_long_extended.csv ready.")
        print(f"  Rows: {len(panel):,}")
        print(f"  Date range: {panel['date'].min().date()} → {panel['date'].max().date()}")
        print(f"  Tickers: {panel['ticker'].nunique()}")
        print(f"  Columns: {list(panel.columns)}")

    except ImportError:
        print("[INFO] blpapi not available on this machine.")
        print("Generating Excel BDH template for manual Bloomberg pull…\n")
        generate_excel_template()
        _print_manual_instructions()
        sys.exit(0)

    except RuntimeError as e:
        print(f"[ERROR] Bloomberg connection failed: {e}")
        print("\nGenerating Excel BDH template as fallback…")
        generate_excel_template()
        _print_manual_instructions()
        sys.exit(1)
